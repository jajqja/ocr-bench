import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_json_to_excel(
    json_file_path, output_excel_path="ocr_evaluation_results.xlsx"
):
    """
    Hàm đọc file JSON kết quả đánh giá mô hình OCR và xuất ra file Excel được định dạng chuyên nghiệp.

    :param json_file_path: Đường dẫn tới file JSON đầu vào.
    :param output_excel_path: Đường dẫn và tên file Excel đầu ra (mặc định: ocr_evaluation_results.xlsx).
    """
    # 1. Đọc dữ liệu từ file JSON
    with open(json_file_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    # 2. Khởi tạo Workbook và thiết lập các kiểu định dạng (Styles)
    wb = openpyxl.Workbook()

    navy_dark = "1B365D"  # Màu xanh navy đậm cho Header
    navy_light = "F0F4F8"  # Màu nền nhẹ xen kẽ giữa các dòng (Zebra striping)
    border_color = "D9D9D9"  # Màu viền xám nhạt

    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11)
    font_body_bold = Font(name="Segoe UI", size=11, bold=True)

    fill_header = PatternFill(
        start_color=navy_dark, end_color=navy_dark, fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color=navy_light, end_color=navy_light, fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    # ----------------------------------------------------
    # TAB 1: Summary Dashboard
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    ws_summary.views.sheetView[0].showGridLines = True  # Hiển thị đường lưới ô

    # Tiêu đề báo cáo
    ws_summary["A1"] = "OCR MODEL EVALUATION REPORT"
    ws_summary["A1"].font = font_title

    # Thông tin Metadata chung
    ws_summary["A3"] = "Dataset:"
    ws_summary["B3"] = json_data.get("dataset", "")
    ws_summary["A4"] = "Model Path:"
    ws_summary["B4"] = json_data.get("model", "")
    ws_summary["A5"] = "Total Samples:"
    ws_summary["B5"] = json_data.get("num_samples", 0)

    for r in range(3, 6):
        ws_summary[f"A{r}"].font = font_body_bold
        ws_summary[f"B{r}"].font = font_body

    # Thông tin thời gian xử lý (Execution Time)
    ws_summary["A7"] = "Performance Metrics"
    ws_summary["A7"].font = font_section

    times_data = json_data.get("times", {})
    ws_summary["A8"] = "Total Time (s)"
    ws_summary["B8"] = times_data.get("total", 0)
    ws_summary["A9"] = "Time per Sample (s)"
    ws_summary["B9"] = times_data.get("per_sample", 0)

    ws_summary["B8"].number_format = "#,##0.02"
    ws_summary["B9"].number_format = "0.0000"
    ws_summary["B8"].alignment = Alignment(horizontal="right")
    ws_summary["B9"].alignment = Alignment(horizontal="right")

    for r in range(8, 10):
        ws_summary[f"A{r}"].font = font_body
        ws_summary[f"B{r}"].font = font_body

    # Bảng số liệu trung bình tổng quan (Mean Metrics)
    ws_summary["A11"] = "Overall Mean Metrics"
    ws_summary["A11"].font = font_section

    headers_metrics = ["Metric", "Value"]
    for col_idx, header in enumerate(headers_metrics, start=1):
        cell = ws_summary.cell(row=12, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    mean_metrics = json_data.get("mean_metrics", {})
    metrics_list = [
        ("Precision", mean_metrics.get("precision", 0)),
        ("Recall", mean_metrics.get("recall", 0)),
        ("F1-Score", mean_metrics.get("f1", 0)),
        ("Page IoU", mean_metrics.get("page_iou", 0)),
    ]

    for idx, (m_name, m_val) in enumerate(metrics_list, start=13):
        c1 = ws_summary.cell(row=idx, column=1, value=m_name)
        c2 = ws_summary.cell(row=idx, column=2, value=m_val)
        c1.font = font_body
        c2.font = font_body_bold
        c2.number_format = "0.00%"
        c2.alignment = Alignment(horizontal="right")
        c1.border = thin_border
        c2.border = thin_border
        if idx % 2 == 0:
            c1.fill = fill_zebra
            c2.fill = fill_zebra

    ws_details = wb.create_sheet(title="Sample Details")
    ws_details.views.sheetView[0].showGridLines = True

    headers_details = ["Sample ID", "Precision", "Recall", "F1-Score", "Page IoU"]
    for col_idx, header in enumerate(headers_details, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    sample_details = json_data.get("sample_details", {})

    for sample_id, metrics in sample_details.items():
        c_id = ws_details.cell(row=row_idx, column=1, value=int(sample_id))
        c_p = ws_details.cell(row=row_idx, column=2, value=metrics.get("precision", 0))
        c_r = ws_details.cell(row=row_idx, column=3, value=metrics.get("recall", 0))
        c_f = ws_details.cell(row=row_idx, column=4, value=metrics.get("f1", 0))
        c_iou = ws_details.cell(row=row_idx, column=5, value=metrics.get("page_iou", 0))

        c_id.alignment = Alignment(horizontal="center")
        for cell in [c_p, c_r, c_f, c_iou]:
            cell.number_format = "0.00%"
            cell.alignment = Alignment(horizontal="right")

        for cell in [c_id, c_p, c_r, c_f, c_iou]:
            cell.font = font_body
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = fill_zebra

        row_idx += 1

    # Thêm hàng tính trung bình (Dùng hàm AVERAGE động của Excel)
    if sample_details:
        c_avg_label = ws_details.cell(row=row_idx, column=1, value="Average")
        c_avg_label.font = font_body_bold
        c_avg_label.alignment = Alignment(horizontal="center")
        c_avg_label.border = thin_border

        for col_idx, col_letter in enumerate(["B", "C", "D", "E"], start=2):
            c_avg = ws_details.cell(
                row=row_idx,
                column=col_idx,
                value=f"=AVERAGE({col_letter}2:{col_letter}{row_idx-1})",
            )
            c_avg.font = font_body_bold
            c_avg.number_format = "0.00%"
            c_avg.alignment = Alignment(horizontal="right")
            c_avg.border = thin_border

    # ----------------------------------------------------
    # Tự động điều chỉnh độ rộng cột tối ưu cho cả 2 tab
    # ----------------------------------------------------
    for sheet in [ws_summary, ws_details]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if val_str.startswith(
                        "="
                    ):  # Giả định độ dài định dạng nếu ô chứa công thức
                        val_str = "100.00%"
                    max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 3. Lưu file Excel kết quả
    wb.save(output_excel_path)
    print(f"Xuất file Excel thành công tại: '{output_excel_path}'")


export_json_to_excel(
    "results/text_detection_results.json", "excels/text_detection.xlsx"
)
