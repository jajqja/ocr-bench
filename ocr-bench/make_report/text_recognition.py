import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_recognition_json_to_excel(
    json_file_path, output_excel_path="recognition_results.xlsx"
):
    """
    Hàm đọc file JSON kết quả đánh giá mô hình Text Recognition và xuất ra file Excel.
    Chỉ thống kê con số (Index, CER, WER), bỏ qua Ground Truth, Prediction và BBox.

    :param json_file_path: Đường dẫn tới file JSON đầu vào.
    :param output_excel_path: Đường dẫn và tên file Excel đầu ra.
    """
    # 1. Đọc dữ liệu từ file JSON
    with open(json_file_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    # 2. Khởi tạo Workbook và thiết lập các kiểu định dạng (Styles)
    wb = openpyxl.Workbook()

    navy_dark = "1B365D"  # Màu xanh navy đậm cho Header
    navy_light = "F0F4F8"  # Màu nền nhẹ xen kẽ giữa các dòng (Zebra striping)
    border_color = "D9D9D9"  # Màu viền xám nhạt

    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11)
    font_body_bold = Font(name="Segoe UI", size=11, bold=True)

    fill_header = PatternFill(
        start_color=navy_dark, end_color=navy_dark, fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color=navy_light, end_color=navy_light, fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    # ----------------------------------------------------
    # TAB 1: Summary Dashboard
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    ws_summary.views.sheetView[0].showGridLines = True

    # Tiêu đề báo cáo
    ws_summary["A1"] = "TEXT RECOGNITION EVALUATION REPORT"
    ws_summary["A1"].font = font_title

    # Thông tin Metadata chung
    ws_summary["A3"] = "Dataset:"
    ws_summary["B3"] = json_data.get("dataset", "")
    ws_summary["A4"] = "Model Path:"
    ws_summary["B4"] = json_data.get("model", "")
    ws_summary["A5"] = "Num Samples:"
    ws_summary["B5"] = json_data.get("num_samples", 0)
    ws_summary["A6"] = "Num Images:"
    ws_summary["B6"] = json_data.get("num_images", 0)

    for r in range(3, 7):
        ws_summary[f"A{r}"].font = font_body_bold
        ws_summary[f"B{r}"].font = font_body

    # Thông tin thời gian xử lý (Inference Time)
    ws_summary["A8"] = "Performance Metrics"
    ws_summary["A8"].font = font_section
    ws_summary["A9"] = "Total Inference Time (s)"
    ws_summary["B9"] = json_data.get("inference_time_total", 0)
    ws_summary["A10"] = "Inference Time per Sample (s)"
    ws_summary["B10"] = json_data.get("inference_time_per_sample", 0)

    ws_summary["B9"].number_format = "#,##0.02"
    ws_summary["B10"].number_format = "0.0000"
    ws_summary["B9"].alignment = Alignment(horizontal="right")
    ws_summary["B10"].alignment = Alignment(horizontal="right")

    for r in range(9, 11):
        ws_summary[f"A{r}"].font = font_body
        ws_summary[f"B{r}"].font = font_body

    # Bảng số liệu trung bình tổng quan (Metrics)
    ws_summary["A12"] = "Overall Metrics"
    ws_summary["A12"].font = font_section

    headers_metrics = ["Metric", "Value"]
    for col_idx, header in enumerate(headers_metrics, start=1):
        cell = ws_summary.cell(row=13, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    metrics_data = json_data.get("metrics", {})
    metrics_list = [
        ("CER (Character Error Rate)", metrics_data.get("cer", 0)),
        ("WER (Word Error Rate)", metrics_data.get("wer", 0)),
        ("Accuracy", metrics_data.get("accuracy", 0)),
    ]

    for idx, (m_name, m_val) in enumerate(metrics_list, start=14):
        c1 = ws_summary.cell(row=idx, column=1, value=m_name)
        c2 = ws_summary.cell(row=idx, column=2, value=m_val)
        c1.font = font_body
        c2.font = font_body_bold
        c2.number_format = "0.00%"
        c2.alignment = Alignment(horizontal="right")
        c1.border = thin_border
        c2.border = thin_border
        if idx % 2 == 1:
            c1.fill = fill_zebra
            c2.fill = fill_zebra

    # ----------------------------------------------------
    # TAB 2: Prediction Details (CHỈ GIỮ LẠI CON SỐ)
    # ----------------------------------------------------
    ws_details = wb.create_sheet(title="Prediction Details")
    ws_details.views.sheetView[0].showGridLines = True

    # Chỉ thống kê: Index, CER, WER
    headers_details = ["Index", "CER", "WER"]
    for col_idx, header in enumerate(headers_details, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    predictions = json_data.get("predictions", [])

    for i, pred_item in enumerate(predictions):
        c_idx = ws_details.cell(row=row_idx, column=1, value=i)
        c_cer = ws_details.cell(row=row_idx, column=2, value=pred_item.get("cer", 0))
        c_wer = ws_details.cell(row=row_idx, column=3, value=pred_item.get("wer", 0))

        c_idx.alignment = Alignment(horizontal="center")
        c_cer.number_format = "0.00%"
        c_wer.number_format = "0.00%"
        c_cer.alignment = Alignment(horizontal="right")
        c_wer.alignment = Alignment(horizontal="right")

        for cell in [c_idx, c_cer, c_wer]:
            cell.font = font_body
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = fill_zebra

        row_idx += 1

    # Thêm hàng tính trung bình (Dùng hàm AVERAGE của Excel cho cột B (CER) và C (WER))
    if predictions:
        c_avg_label = ws_details.cell(row=row_idx, column=1, value="Average")
        c_avg_label.font = font_body_bold
        c_avg_label.alignment = Alignment(horizontal="center")
        c_avg_label.border = thin_border

        # Điền công thức AVERAGE cho cột B (CER) và C (WER) tương ứng cấu trúc mới
        for col_idx, col_letter in enumerate(["B", "C"], start=2):
            c_avg = ws_details.cell(
                row=row_idx,
                column=col_idx,
                value=f"=AVERAGE({col_letter}2:{col_letter}{row_idx-1})",
            )
            c_avg.font = font_body_bold
            c_avg.number_format = "0.00%"
            c_avg.alignment = Alignment(horizontal="right")
            c_avg.border = thin_border

    # ----------------------------------------------------
    # Tự động điều chỉnh độ rộng cột tối ưu cho cả 2 tab
    # ----------------------------------------------------
    for sheet in [ws_summary, ws_details]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if val_str.startswith(
                        "="
                    ):  # Tránh lấy độ dài của chuỗi công thức Excel
                        val_str = "100.00%"
                    max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 3. Lưu file Excel kết quả
    wb.save(output_excel_path)
    print(f"Xuất file Excel thành công tại: '{output_excel_path}'")


# --- HƯỚNG DẪN SỬ DỤNG VÍ DỤ ---
export_recognition_json_to_excel(
    "results/text_recognition_results.json", "excels/text_recognition.xlsx"
)
